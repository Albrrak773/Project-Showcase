import openpyxl
import csv


wb = openpyxl.load_workbook("Graduation_Projects_data_2025.xlsx")
sheet = wb.active

with open('Categorized_Tags.csv', 'r', encoding='utf-8') as tags_file:
    reader = csv.reader(tags_file, delimiter='\n')

    i = 1
    for row in reader:
        sheet.cell(row=i, column=10).value = row[0]
        i += 1

with open("Cleaned_Names.csv", "r",encoding='utf-8') as names_file:
    reader = csv.reader(names_file, delimiter='\n')

    i = 1
    for row in reader:
        sheet.cell(row=i, column=7).value = row[0]
        i += 1

wb.save("Graduation_Projects_data_2025.xlsx")